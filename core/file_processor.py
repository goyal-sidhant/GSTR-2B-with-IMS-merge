# core/file_processor.py
"""
Main file processing logic for GST Processing Tool v5.0

This is the central orchestrator that:
- Coordinates all processing operations
- Manages the processing workflow
- Calls other modules (validators, excel_handler, report_generator)
- Emits progress updates for UI
"""

import shutil
import time
from pathlib import Path
from typing import List, Dict, Callable, Optional, Tuple, Any

from core.models import (
    ProcessingResult, ValidationResult, ClientInfo,
    SummaryData, ReportData, ComparisonData, PreviewData,
    ProcessingSummary, ProgressInfo
)
from core.excel_handler import ExcelHandler
from core.validators import FileValidator
from core.report_generator import ReportGenerator
from utils.file_utils import (
    check_file_month, extract_client_info, find_extra_files,
    get_best_file, create_output_folders, organize_files_by_client,
    get_unique_clients
)
from utils.date_utils import format_timestamp
from utils.constants import StatusType, ErrorCode, CategoryType
from utils.logger import get_logger


class FileProcessor:
    """
    Main processor that orchestrates all GST file operations.
    
    This class is the "brain" of the application - it coordinates:
    - File validation
    - Client scanning
    - File processing (merge/create/copy)
    - Report generation
    - Progress reporting
    """
    
    def __init__(self):
        self.logger = get_logger()
        self.excel_handler = ExcelHandler()
        self.validator = FileValidator()
        self.report_generator = ReportGenerator()
        
        # Progress callback
        self._progress_callback: Optional[Callable[[ProgressInfo], None]] = None
    
    def set_progress_callback(self, callback: Callable[[ProgressInfo], None]) -> None:
        """
        Set callback function for progress updates.
        
        Args:
            callback: Function that receives ProgressInfo
        """
        self._progress_callback = callback
    
    def _emit_progress(self, current: int, total: int, client_name: str, message: str = "") -> None:
        """Emit progress update to UI."""
        if self._progress_callback:
            progress = ProgressInfo(
                current=current,
                total=total,
                client_name=client_name,
                message=message
            )
            self._progress_callback(progress)
    
    def scan_folder(self, input_path: str, month: str, year: str) -> Tuple[List[ClientInfo], List[str]]:
        """
        Scan folder and return client list and extra files.
        
        Args:
            input_path: Path to input folder
            month: Selected month
            year: Selected year
        
        Returns:
            Tuple of (client_list, extra_files)
        """
        self.logger.info(f"Scanning folder: {input_path}")
        
        path = Path(input_path)
        clients = self.validator.scan_clients(path, month, year)
        extra_files = find_extra_files(path, month, year)
        
        self.logger.info(f"Found {len(clients)} clients, {len(extra_files)} extra files")
        
        return clients, extra_files
    
    def get_preview(self, clients: List[ClientInfo], extra_files: List[str]) -> PreviewData:
        """
        Get preview data for UI display.
        
        Args:
            clients: List of client info
            extra_files: List of extra file names
        
        Returns:
            PreviewData for display
        """
        return self.validator.get_preview_data(clients, extra_files)
    
    def validate_folder(self, input_path: str, month: str, year: str,
                       total_clients: int = 0, not_generated: int = 0) -> ValidationResult:
        """
        Validate input folder.
        
        Args:
            input_path: Path to folder
            month: Selected month
            year: Selected year
            total_clients: Expected total
            not_generated: Not generated count
        
        Returns:
            ValidationResult
        """
        return self.validator.validate_folder(
            Path(input_path), month, year, total_clients, not_generated
        )
    
    def auto_detect_folder_info(self, folder_path: str) -> Tuple[Optional[str], Optional[str], Optional[int]]:
        """
        Auto-detect month, year, client count from folder.
        
        Args:
            folder_path: Path to folder
        
        Returns:
            Tuple of (month, year, client_count)
        """
        return self.validator.auto_detect_folder_info(folder_path)
    
    def process_files(self, input_path: str, month: str, year: str,
                     selected_clients: List[ClientInfo],
                     comparison_files: List[str] = None) -> ProcessingSummary:
        """
        Main processing function.
        
        This is called when user clicks "Start Processing".
        
        Args:
            input_path: Path to input folder
            month: Processing month
            year: Processing year
            selected_clients: List of clients to process (only selected ones)
            comparison_files: Optional list of comparison report paths
        
        Returns:
            ProcessingSummary with results
        """
        start_time = time.time()
        
        self.logger.info("=" * 50)
        self.logger.info(f"Starting processing for {month} {year}")
        self.logger.info(f"Clients to process: {len(selected_clients)}")
        self.logger.separator()
        
        input_path = Path(input_path)
        month_year = f"{month} {year}"
        timestamp = format_timestamp()
        
        # Create output folders
        output_main, processed_folder, created_folder = create_output_folders(
            input_path.parent, month_year, timestamp
        )
        
        self.logger.info(f"Output folder: {output_main}")
        
        # Report path
        report_path = output_main / f"GSTR-2B Processing Report {month_year}_{timestamp}.xlsx"
        
        # Get all files for processing
        all_gstr2b = [f for f in input_path.glob("GSTR2B-*.xlsx")
                     if check_file_month(f.stem, month, year)]
        all_ims = list(input_path.glob("ImsReco-*.xlsx"))
        
        # Organize by client
        gstr_by_client = organize_files_by_client(all_gstr2b, "GSTR2B")
        
        ims_by_client = {}
        for f in all_ims:
            name, state = extract_client_info(f.stem, "IMS")
            if name and state:
                ims_by_client[f"{name}-{state}"] = f
        
        # Find template for creating new files
        template_file = self.excel_handler.find_template_gstr2b(input_path, month, year)
        
        # Extra files
        extra_files = find_extra_files(input_path, month, year)
        
        # Process each selected client
        summary = ProcessingSummary()
        all_client_results = []
        total = len(selected_clients)
        
        for i, client in enumerate(selected_clients):
            # Emit progress
            self._emit_progress(i + 1, total, client.name, f"Processing {client.name}")
            
            # Process this client
            result_client = self._process_single_client(
                client,
                gstr_by_client,
                ims_by_client,
                processed_folder,
                created_folder,
                month, year,
                template_file
            )
            
            all_client_results.append(result_client)
            
            # Update summary counts
            if result_client.issue_type == StatusType.NONE.value:
                summary.merged += 1
                self.logger.merged(client.name)
            elif result_client.issue_type == StatusType.CREATED.value:
                summary.created += 1
                self.logger.created(client.name)
            elif result_client.issue_type == StatusType.WARNING.value:
                summary.copied += 1
                self.logger.copied(client.name)
            else:
                summary.errors += 1
                self.logger.failed(client.name, result_client.status)
        
        # Parse comparison files
        comparison_data = self._parse_comparison_files(comparison_files or [])
        
        # Add current period data to comparison
        current_summary = SummaryData(
            clients_with_gstr2b=len(gstr_by_client),
            clients_with_ims=len(ims_by_client),
            clients_with_both=sum(1 for c in all_client_results 
                                 if c.has_gstr2b and c.has_ims),
            clients_gstr_only=sum(1 for c in all_client_results 
                                 if c.has_gstr2b and not c.has_ims),
            clients_ims_only=sum(1 for c in all_client_results 
                                if c.has_ims and not c.has_gstr2b),
            files_processed=summary.merged,
            files_created=summary.created,
            files_with_errors=summary.errors,
            extra_files_count=len(extra_files)
        )
        
        current_comparison = {
            'month': month,
            'year': year,
            'period': month_year,
            'file_path': str(report_path),
            'clients_with_gstr2b': current_summary.clients_with_gstr2b,
            'clients_with_ims': current_summary.clients_with_ims,
            'clients_with_both': current_summary.clients_with_both,
            'files_processed': current_summary.files_processed,
            'files_created': current_summary.files_created,
            'files_with_errors': current_summary.files_with_errors,
            'clients': [{'name': c.name, 'state': c.state, 'issue_type': c.issue_type}
                       for c in all_client_results],
            'error_patterns': [c.issue_type for c in all_client_results
                              if c.issue_type not in ['None', 'Created']]
        }
        comparison_data.append(current_comparison)
        
        # Generate report
        self._emit_progress(total, total, "", "Generating report...")
        self.logger.info("Generating Excel report...")
        
        report_data = ReportData(
            client_data=all_client_results,
            summary=current_summary,
            extra_files=extra_files,
            month_year=month_year,
            comparison_data=comparison_data
        )
        
        self.report_generator.create_report(report_path, report_data)
        
        # Calculate duration
        duration = time.time() - start_time
        
        # Complete summary
        summary.output_folder = output_main
        summary.report_path = report_path
        summary.duration_seconds = duration
        
        self.logger.separator()
        self.logger.info(f"Processing complete in {summary.duration_text}")
        self.logger.info(f"Merged: {summary.merged}, Created: {summary.created}, "
                        f"Copied: {summary.copied}, Errors: {summary.errors}")
        self.logger.info(f"Report: {report_path.name}")
        self.logger.info("=" * 50)
        
        return summary
    
    def _process_single_client(self, client: ClientInfo,
                               gstr_by_client: Dict[str, List[Path]],
                               ims_by_client: Dict[str, Path],
                               processed_folder: Path,
                               created_folder: Path,
                               month: str, year: str,
                               template_file: Optional[Path]) -> ClientInfo:
        """
        Process a single client.
        
        Args:
            client: ClientInfo object
            gstr_by_client: GSTR-2B files by client key
            ims_by_client: IMS files by client key
            processed_folder: Folder for merged files
            created_folder: Folder for created files
            month: Processing month
            year: Processing year
            template_file: Template GSTR-2B file
        
        Returns:
            Updated ClientInfo with processing results
        """
        client_key = client.client_key
        
        try:
            if client.category == CategoryType.MERGE.value:
                # Merge GSTR-2B with IMS
                gstr_file = get_best_file(gstr_by_client.get(client_key, []))
                ims_file = ims_by_client.get(client_key)
                
                if gstr_file and ims_file:
                    output_file = processed_folder / gstr_file.name
                    result = self.excel_handler.merge_ims_into_gstr2b(
                        gstr_file, ims_file, output_file
                    )
                    
                    if result.success:
                        client.status = "Successfully merged GSTR-2B with IMS"
                        client.issue_type = StatusType.NONE.value
                    else:
                        client.status = result.message
                        client.issue_type = StatusType.ERROR.value
                        client.error_code = result.error_code or ""
                else:
                    client.status = "Files not found for merge"
                    client.issue_type = StatusType.ERROR.value
                    client.error_code = ErrorCode.FILE_NOT_FOUND.value
            
            elif client.category == CategoryType.CREATE.value:
                # Create GSTR-2B from IMS
                ims_file = ims_by_client.get(client_key)
                
                if ims_file:
                    result = self.excel_handler.create_gstr2b_from_ims(
                        ims_file, created_folder, month, year, template_file
                    )
                    
                    if result.success:
                        client.status = "Created GSTR-2B from IMS Reco"
                        client.issue_type = StatusType.CREATED.value
                    else:
                        client.status = result.message
                        client.issue_type = StatusType.ERROR.value
                        client.error_code = result.error_code or ""
                else:
                    client.status = "IMS file not found"
                    client.issue_type = StatusType.ERROR.value
                    client.error_code = ErrorCode.FILE_NOT_FOUND.value
            
            elif client.category == CategoryType.COPY.value:
                # Copy GSTR-2B only (no IMS)
                gstr_file = get_best_file(gstr_by_client.get(client_key, []))
                
                if gstr_file:
                    result = self.excel_handler.copy_file(gstr_file, processed_folder)
                    
                    if result.success:
                        client.status = "Copied GSTR-2B (no IMS Reco available)"
                        client.issue_type = StatusType.WARNING.value
                        client.error_code = ErrorCode.FILE_NOT_FOUND.value
                    else:
                        client.status = result.message
                        client.issue_type = StatusType.ERROR.value
                        client.error_code = result.error_code or ""
                else:
                    client.status = "GSTR-2B file not found"
                    client.issue_type = StatusType.ERROR.value
                    client.error_code = ErrorCode.FILE_NOT_FOUND.value
        
        except Exception as e:
            client.status = f"Processing error: {str(e)}"
            client.issue_type = StatusType.ERROR.value
            client.error_code = ErrorCode.GENERAL_ERROR.value
        
        return client
    
    def _parse_comparison_files(self, file_paths: List[str]) -> List[Dict[str, Any]]:
        """
        Parse comparison report files.
        
        Args:
            file_paths: List of file paths
        
        Returns:
            List of parsed data dictionaries
        """
        comparison_data = []
        
        for file_path in file_paths:
            try:
                from openpyxl import load_workbook
                from utils.file_utils import extract_month_year_from_report
                
                path = Path(file_path)
                if not path.exists():
                    continue
                
                wb = load_workbook(path, data_only=True)
                month, year = extract_month_year_from_report(path)
                
                # Parse Summary sheet
                data = {
                    'month': month or '',
                    'year': year or '',
                    'period': f"{month} {year}" if month and year else "Unknown",
                    'file_path': str(path),
                    'clients_with_gstr2b': 0,
                    'clients_with_ims': 0,
                    'clients_with_both': 0,
                    'files_processed': 0,
                    'files_created': 0,
                    'files_with_errors': 0,
                    'clients': [],
                    'error_patterns': []
                }
                
                if 'Summary' in wb.sheetnames:
                    sheet = wb['Summary']
                    for row in range(1, sheet.max_row + 1):
                        label = str(sheet.cell(row=row, column=1).value or '').lower()
                        value = sheet.cell(row=row, column=2).value
                        
                        if 'gstr-2b files' in label and 'total' in label:
                            data['clients_with_gstr2b'] = int(value) if value else 0
                        elif 'ims reco files' in label and 'total' in label:
                            data['clients_with_ims'] = int(value) if value else 0
                        elif 'both files' in label:
                            data['clients_with_both'] = int(value) if value else 0
                        elif 'successfully processed' in label:
                            data['files_processed'] = int(value) if value else 0
                        elif 'created' in label:
                            data['files_created'] = int(value) if value else 0
                        elif 'errors' in label:
                            data['files_with_errors'] = int(value) if value else 0
                
                # Parse Detailed Report for clients
                if 'Detailed Report' in wb.sheetnames:
                    sheet = wb['Detailed Report']
                    for row in range(2, sheet.max_row + 1):
                        name = sheet.cell(row=row, column=1).value
                        state = sheet.cell(row=row, column=2).value
                        issue = sheet.cell(row=row, column=6).value
                        
                        if name:
                            data['clients'].append({
                                'name': str(name),
                                'state': str(state) if state else '',
                                'issue_type': str(issue) if issue else 'None'
                            })
                            if issue and str(issue) not in ['None', 'Created']:
                                data['error_patterns'].append(str(issue))
                
                comparison_data.append(data)
                self.logger.info(f"Parsed comparison file: {path.name}")
                
            except Exception as e:
                self.logger.warning(f"Could not parse {file_path}: {e}")
        
        return comparison_data
