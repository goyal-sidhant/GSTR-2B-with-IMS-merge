# core/report_generator.py
"""
Report generation for GST Processing Tool v5.0

Creates detailed Excel reports with multiple sheets.
NOTE: Charts have been removed for simplicity - only data tables.
"""

from pathlib import Path
from typing import List, Dict, Any
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

from core.models import ReportData, ClientInfo, SummaryData
from utils.constants import ExcelColors, SheetNames
from utils.logger import get_logger


class ReportGenerator:
    """
    Generates Excel reports with processing results.
    
    Creates multiple sheets with different views of the data.
    Charts have been removed - only data tables are generated.
    """
    
    def __init__(self):
        self.logger = get_logger()
        self.styles = self._create_styles()
    
    def _create_styles(self) -> Dict[str, Any]:
        """Create reusable styles for the report."""
        return {
            'error_fill': PatternFill(
                start_color=ExcelColors.ERROR_RED,
                end_color=ExcelColors.ERROR_RED,
                fill_type="solid"
            ),
            'warning_fill': PatternFill(
                start_color=ExcelColors.WARNING_ORANGE,
                end_color=ExcelColors.WARNING_ORANGE,
                fill_type="solid"
            ),
            'created_fill': PatternFill(
                start_color=ExcelColors.CREATED_BLUE,
                end_color=ExcelColors.CREATED_BLUE,
                fill_type="solid"
            ),
            'success_fill': PatternFill(
                start_color=ExcelColors.SUCCESS_GREEN,
                end_color=ExcelColors.SUCCESS_GREEN,
                fill_type="solid"
            ),
            'header_fill': PatternFill(
                start_color=ExcelColors.HEADER_BLUE,
                end_color=ExcelColors.HEADER_BLUE,
                fill_type="solid"
            ),
            'header_font': Font(bold=True, color="FFFFFF"),
            'bold_font': Font(bold=True),
            'error_font': Font(color="FFFFFF", bold=True),
            'warning_font': Font(color="000000", bold=True),
            'created_font': Font(color="000000", bold=True)
        }
    
    def _set_tab_color(self, sheet, color: str) -> None:
        """Set worksheet tab color."""
        sheet.sheet_properties.tabColor = color
    
    def create_report(self, report_path: Path, report_data: ReportData) -> None:
        """
        Create the complete Excel report.
        
        Args:
            report_path: Path to save the report
            report_data: All data for the report
        """
        self.logger.info(f"Creating report: {report_path.name}")
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create core sheets
        self._create_summary_sheet(wb, report_data)
        self._create_detailed_report_sheet(wb, report_data)
        self._create_error_details_sheet(wb, report_data)
        self._create_extra_files_sheet(wb, report_data)
        
        # Create comparison sheets if data available
        if len(report_data.comparison_data) > 1:
            self._create_monthly_trends_sheet(wb, report_data)
            self._create_error_analytics_sheet(wb, report_data)
            self._create_efficiency_dashboard_sheet(wb, report_data)
            self._create_client_journey_sheet(wb, report_data)
            self._create_executive_dashboard_sheet(wb, report_data)
            self._create_action_items_sheet(wb, report_data)
        
        # Save workbook
        wb.save(report_path)
        self.logger.info("Report created successfully")
    
    def _create_summary_sheet(self, wb: Workbook, report_data: ReportData) -> None:
        """Create Summary sheet."""
        sheet = wb.create_sheet(SheetNames.SUMMARY)
        self._set_tab_color(sheet, ExcelColors.TAB_BLUE)
        
        summary = report_data.summary
        
        # Summary data rows
        rows = [
            ["GST Processing Report", f"{report_data.month_year}"],
            ["", ""],
            ["Total Clients with GSTR-2B Files", summary.clients_with_gstr2b],
            ["Total Clients with IMS Reco Files", summary.clients_with_ims],
            ["Clients with Both Files", summary.clients_with_both],
            ["Clients with GSTR-2B Only", summary.clients_gstr_only],
            ["Clients with IMS Reco Only", summary.clients_ims_only],
            ["", ""],
            ["Files Successfully Processed (Merged)", summary.files_processed],
            ["GSTR-2B Files Created from IMS", summary.files_created],
            ["Files with Errors/Issues", summary.files_with_errors],
            ["Extra Files Found", summary.extra_files_count],
            ["", ""],
            ["Report Generated", report_data.processing_timestamp.strftime("%d/%m/%Y %I:%M:%S %p")]
        ]
        
        for row_num, (label, value) in enumerate(rows, 1):
            cell_a = sheet.cell(row=row_num, column=1, value=label)
            cell_b = sheet.cell(row=row_num, column=2, value=value)
            
            if row_num == 1:
                cell_a.font = Font(bold=True, size=14)
            elif label:
                cell_a.font = self.styles['bold_font']
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 25
    
    def _create_detailed_report_sheet(self, wb: Workbook, report_data: ReportData) -> None:
        """Create Detailed Report sheet with all clients."""
        sheet = wb.create_sheet(SheetNames.DETAILED_REPORT)
        self._set_tab_color(sheet, ExcelColors.TAB_BLUE)
        
        # Headers
        headers = ["Client Name", "State", "GSTR-2B", "IMS Reco", 
                  "Processing Status", "Issue Type"]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = self.styles['header_fill']
            cell.font = self.styles['header_font']
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_num, client in enumerate(report_data.client_data, 2):
            data = [
                client.name,
                client.state,
                "Yes" if client.has_gstr2b else "No",
                "Yes" if client.has_ims else "No",
                client.status,
                client.issue_type
            ]
            
            for col, value in enumerate(data, 1):
                cell = sheet.cell(row=row_num, column=col, value=value)
                
                # Color coding
                if client.issue_type == "Error":
                    cell.fill = self.styles['error_fill']
                    cell.font = self.styles['error_font']
                elif client.issue_type == "Warning":
                    cell.fill = self.styles['warning_fill']
                    cell.font = self.styles['warning_font']
                elif client.issue_type == "Created":
                    cell.fill = self.styles['created_fill']
                    cell.font = self.styles['created_font']
        
        # Adjust column widths
        widths = [30, 10, 12, 12, 40, 15]
        for col, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(col)].width = width
    
    def _create_error_details_sheet(self, wb: Workbook, report_data: ReportData) -> None:
        """Create Error Details sheet."""
        error_clients = [c for c in report_data.client_data 
                        if c.issue_type in ["Error", "Warning"]]
        
        if not error_clients:
            return
        
        sheet = wb.create_sheet(SheetNames.ERROR_DETAILS)
        self._set_tab_color(sheet, ExcelColors.TAB_BLUE)
        
        # Headers
        headers = ["Client Name", "State", "Issue Type", "Description"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = self.styles['header_fill']
            cell.font = self.styles['header_font']
        
        # Data rows
        for row_num, client in enumerate(error_clients, 2):
            sheet.cell(row=row_num, column=1, value=client.name)
            sheet.cell(row=row_num, column=2, value=client.state)
            sheet.cell(row=row_num, column=3, value=client.issue_type)
            sheet.cell(row=row_num, column=4, value=client.status)
            
            # Color coding
            for col in range(1, 5):
                cell = sheet.cell(row=row_num, column=col)
                if client.issue_type == "Error":
                    cell.fill = self.styles['error_fill']
                    cell.font = self.styles['error_font']
                else:
                    cell.fill = self.styles['warning_fill']
                    cell.font = self.styles['warning_font']
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 50
    
    def _create_extra_files_sheet(self, wb: Workbook, report_data: ReportData) -> None:
        """Create Extra Files sheet."""
        if not report_data.extra_files:
            return
        
        sheet = wb.create_sheet(SheetNames.EXTRA_FILES)
        self._set_tab_color(sheet, ExcelColors.TAB_BLUE)
        
        # Header
        cell = sheet.cell(row=1, column=1, value="Extra Files Found (Not Processed)")
        cell.fill = self.styles['header_fill']
        cell.font = self.styles['header_font']
        
        # List files
        for row_num, filename in enumerate(report_data.extra_files, 2):
            cell = sheet.cell(row=row_num, column=1, value=filename)
            cell.fill = self.styles['warning_fill']
        
        sheet.column_dimensions['A'].width = 60
    
    def _create_monthly_trends_sheet(self, wb: Workbook, report_data: ReportData) -> None:
        """Create Monthly Trends sheet (data only, no charts)."""
        sheet = wb.create_sheet(SheetNames.MONTHLY_TRENDS)
        self._set_tab_color(sheet, ExcelColors.TAB_GREEN)
        
        # Sort by period
        sorted_data = sorted(
            report_data.comparison_data,
            key=lambda x: (x.get('year', ''), x.get('month', ''))
        )
        
        # Headers
        headers = ["Period", "GSTR-2B Clients", "IMS Clients", 
                  "Files Processed", "Files Created", "Errors", "Growth %"]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = self.styles['header_fill']
            cell.font = self.styles['header_font']
        
        # Data rows
        prev_total = 0
        for row_num, data in enumerate(sorted_data, 2):
            current_total = data.get('clients_with_gstr2b', 0)
            
            # Calculate growth
            if prev_total > 0:
                growth = ((current_total - prev_total) / prev_total) * 100
                growth_str = f"{growth:+.1f}%"
            else:
                growth_str = "N/A"
            
            row_data = [
                data.get('period', 'Unknown'),
                current_total,
                data.get('clients_with_ims', 0),
                data.get('files_processed', 0),
                data.get('files_created', 0),
                data.get('files_with_errors', 0),
                growth_str
            ]
            
            for col, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col, value=value)
            
            prev_total = current_total
        
        # Adjust widths
        for col in range(1, 8):
            sheet.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_error_analytics_sheet(self, wb: Workbook, report_data: ReportData) -> None:
        """Create Error Analytics sheet (data only, no charts)."""
        sheet = wb.create_sheet(SheetNames.ERROR_ANALYTICS)
        self._set_tab_color(sheet, ExcelColors.TAB_GREEN)
        
        # Collect all errors
        all_errors = []
        error_by_month = {}
        
        for data in report_data.comparison_data:
            period = data.get('period', 'Unknown')
            errors = data.get('error_patterns', [])
            error_by_month[period] = len(errors)
            all_errors.extend(errors)
        
        error_counter = Counter(all_errors)
        
        # Title
        sheet.cell(row=1, column=1, value="Error Analytics").font = Font(bold=True, size=14)
        
        # Most common errors section
        sheet.cell(row=3, column=1, value="Most Common Error Types").font = self.styles['bold_font']
        sheet.cell(row=4, column=1, value="Error Type").font = self.styles['bold_font']
        sheet.cell(row=4, column=2, value="Count").font = self.styles['bold_font']
        
        row = 5
        for error_type, count in error_counter.most_common(10):
            sheet.cell(row=row, column=1, value=error_type)
            sheet.cell(row=row, column=2, value=count)
            row += 1
        
        # Errors by month section
        row += 2
        sheet.cell(row=row, column=1, value="Error Count by Period").font = self.styles['bold_font']
        row += 1
        sheet.cell(row=row, column=1, value="Period").font = self.styles['bold_font']
        sheet.cell(row=row, column=2, value="Errors").font = self.styles['bold_font']
        row += 1
        
        for period, count in error_by_month.items():
            sheet.cell(row=row, column=1, value=period)
            sheet.cell(row=row, column=2, value=count)
            row += 1
        
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 15
    
    def _create_efficiency_dashboard_sheet(self, wb: Workbook, report_data: ReportData) -> None:
        """Create Efficiency Dashboard sheet (data only, no charts)."""
        sheet = wb.create_sheet(SheetNames.EFFICIENCY_DASHBOARD)
        self._set_tab_color(sheet, ExcelColors.TAB_GREEN)
        
        # Headers
        headers = ["Period", "Total Clients", "Success Rate", "Error Rate", "Files Completed"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = self.styles['header_fill']
            cell.font = self.styles['header_font']
        
        # Calculate metrics for each period
        row = 2
        for data in report_data.comparison_data:
            total = max(data.get('clients_with_gstr2b', 1), 1)
            completed = data.get('files_processed', 0) + data.get('files_created', 0)
            errors = data.get('files_with_errors', 0)
            
            success_rate = (completed / total) * 100
            error_rate = (errors / total) * 100
            
            sheet.cell(row=row, column=1, value=data.get('period', 'Unknown'))
            sheet.cell(row=row, column=2, value=total)
            sheet.cell(row=row, column=3, value=f"{success_rate:.1f}%")
            sheet.cell(row=row, column=4, value=f"{error_rate:.1f}%")
            sheet.cell(row=row, column=5, value=completed)
            row += 1
        
        for col in range(1, 6):
            sheet.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_client_journey_sheet(self, wb: Workbook, report_data: ReportData) -> None:
        """Create Client Journey Tracker sheet."""
        sheet = wb.create_sheet(SheetNames.CLIENT_JOURNEY)
        self._set_tab_color(sheet, ExcelColors.TAB_GREEN)
        
        # Collect all clients and their status per period
        all_clients = set()
        client_status = defaultdict(dict)
        periods = []
        
        for data in report_data.comparison_data:
            period = data.get('period', 'Unknown')
            periods.append(period)
            
            for client in data.get('clients', []):
                client_key = f"{client.get('name', '')}-{client.get('state', '')}"
                all_clients.add(client_key)
                client_status[client_key][period] = client.get('issue_type', 'None')
        
        # Headers
        cell = sheet.cell(row=1, column=1, value="Client")
        cell.fill = self.styles['header_fill']
        cell.font = self.styles['header_font']
        
        for col, period in enumerate(periods, 2):
            cell = sheet.cell(row=1, column=col, value=period)
            cell.fill = self.styles['header_fill']
            cell.font = self.styles['header_font']
        
        # Data rows
        for row_num, client in enumerate(sorted(all_clients), 2):
            sheet.cell(row=row_num, column=1, value=client)
            
            for col, period in enumerate(periods, 2):
                status = client_status[client].get(period, "Not Present")
                
                # Convert to display text
                if status == "None":
                    display = "✓ Processed"
                elif status == "Error":
                    display = "✗ Error"
                elif status == "Warning":
                    display = "⚠ Warning"
                elif status == "Created":
                    display = "✓ Created"
                else:
                    display = "- Not Present"
                
                cell = sheet.cell(row=row_num, column=col, value=display)
                
                # Color coding
                if status == "Error":
                    cell.fill = self.styles['error_fill']
                elif status == "Warning":
                    cell.fill = self.styles['warning_fill']
                elif status == "Created":
                    cell.fill = self.styles['created_fill']
        
        sheet.column_dimensions['A'].width = 35
        for col in range(2, len(periods) + 2):
            sheet.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_executive_dashboard_sheet(self, wb: Workbook, report_data: ReportData) -> None:
        """Create Executive Dashboard sheet."""
        sheet = wb.create_sheet(SheetNames.EXECUTIVE_DASHBOARD)
        self._set_tab_color(sheet, ExcelColors.TAB_ORANGE)
        
        latest = report_data.comparison_data[-1] if report_data.comparison_data else {}
        prev = report_data.comparison_data[-2] if len(report_data.comparison_data) > 1 else {}
        
        # Calculate metrics
        total = max(latest.get('clients_with_gstr2b', 1), 1)
        completed = latest.get('files_processed', 0) + latest.get('files_created', 0)
        success_rate = (completed / total) * 100
        error_rate = (latest.get('files_with_errors', 0) / total) * 100
        
        # Growth calculation
        if prev:
            growth = latest.get('clients_with_gstr2b', 0) - prev.get('clients_with_gstr2b', 0)
            error_change = prev.get('files_with_errors', 0) - latest.get('files_with_errors', 0)
        else:
            growth = 0
            error_change = 0
        
        # Title
        sheet.cell(row=1, column=1, value="EXECUTIVE DASHBOARD").font = Font(bold=True, size=14)
        
        rows = [
            ["", ""],
            ["Current Period", latest.get('period', 'Unknown')],
            ["Total Clients (GSTR-2B)", latest.get('clients_with_gstr2b', 0)],
            ["Processing Success Rate", f"{success_rate:.1f}%"],
            ["Files Created from IMS", latest.get('files_created', 0)],
            ["Error Rate", f"{error_rate:.1f}%"],
            ["", ""],
            ["MONTH-OVER-MONTH CHANGES", ""],
            ["Client Growth", f"{growth:+d}" if prev else "N/A"],
            ["Error Reduction", f"{error_change:+d}" if prev else "N/A"],
        ]
        
        for row_num, (label, value) in enumerate(rows, 2):
            cell_a = sheet.cell(row=row_num, column=1, value=label)
            cell_b = sheet.cell(row=row_num, column=2, value=value)
            
            if "MONTH-OVER-MONTH" in label or row_num == 1:
                cell_a.font = Font(bold=True)
            elif label:
                cell_a.font = self.styles['bold_font']
        
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 20
    
    def _create_action_items_sheet(self, wb: Workbook, report_data: ReportData) -> None:
        """Create Action Items sheet."""
        sheet = wb.create_sheet(SheetNames.ACTION_ITEMS)
        self._set_tab_color(sheet, ExcelColors.TAB_ORANGE)
        
        sheet.cell(row=1, column=1, value="ACTION ITEMS & RECOMMENDATIONS").font = \
            Font(bold=True, size=14)
        
        # Generate action items based on data
        actions = []
        
        latest = report_data.comparison_data[-1] if report_data.comparison_data else {}
        prev = report_data.comparison_data[-2] if len(report_data.comparison_data) > 1 else {}
        
        # Client growth analysis
        if prev:
            growth = latest.get('clients_with_gstr2b', 0) - prev.get('clients_with_gstr2b', 0)
            if growth > 0:
                actions.append(f"✓ Positive: Client base grew by {growth} clients")
            elif growth < 0:
                actions.append(f"⚠ Attention: Client base decreased by {abs(growth)} clients")
        
        # Error rate analysis
        total = max(latest.get('clients_with_gstr2b', 1), 1)
        error_rate = (latest.get('files_with_errors', 0) / total) * 100
        
        if error_rate > 10:
            actions.append(f"🔴 High Priority: Error rate is {error_rate:.1f}% - Investigate issues")
        elif error_rate > 5:
            actions.append(f"🟡 Medium Priority: Error rate is {error_rate:.1f}% - Monitor closely")
        else:
            actions.append(f"✓ Good: Error rate is {error_rate:.1f}% - Within acceptable range")
        
        # Files created
        created = latest.get('files_created', 0)
        if created > 0:
            actions.append(f"📋 Process: {created} GSTR-2B files created from IMS - Verify accuracy")
        
        # Write actions
        for row_num, action in enumerate(actions, 3):
            sheet.cell(row=row_num, column=1, value=action)
        
        sheet.column_dimensions['A'].width = 80
