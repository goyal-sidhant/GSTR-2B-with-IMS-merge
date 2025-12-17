# core/excel_handler.py
"""
Excel file handling operations for GST Processing Tool v5.0

This module handles:
- Merging IMS sheet into GSTR-2B files
- Creating new GSTR-2B files from IMS Reco
- Copying sheets with formatting
- All direct Excel file operations
"""

import datetime
from pathlib import Path
from typing import Optional, Tuple, Any

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from core.models import ProcessingResult
from utils.file_utils import extract_client_info
from utils.logger import get_logger


class ExcelHandler:
    """
    Handles all Excel file operations.
    """
    
    def __init__(self):
        self.logger = get_logger()
    
    def copy_sheet_with_formatting(self, source_wb: Workbook, source_sheet_name: str,
                                   target_wb: Workbook, target_sheet_name: str) -> Any:
        """
        Copy a worksheet from one workbook to another, preserving all formatting.
        
        Args:
            source_wb: Source workbook
            source_sheet_name: Name of sheet to copy
            target_wb: Target workbook
            target_sheet_name: Name for the new sheet
        
        Returns:
            The new sheet in the target workbook
        """
        # Get source sheet
        source_sheet = source_wb[source_sheet_name]
        
        # Remove existing sheet if it exists
        if target_sheet_name in target_wb.sheetnames:
            target_wb.remove(target_wb[target_sheet_name])
        
        # Create new sheet
        target_sheet = target_wb.create_sheet(target_sheet_name)
        
        # Copy column dimensions (widths)
        for col in range(1, source_sheet.max_column + 1):
            col_letter = get_column_letter(col)
            if col_letter in source_sheet.column_dimensions:
                target_sheet.column_dimensions[col_letter].width = \
                    source_sheet.column_dimensions[col_letter].width
        
        # Copy row dimensions (heights)
        for row in range(1, source_sheet.max_row + 1):
            if row in source_sheet.row_dimensions:
                target_sheet.row_dimensions[row].height = \
                    source_sheet.row_dimensions[row].height
        
        # Copy merged cells
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))
        
        # Copy cell by cell
        for row in source_sheet.rows:
            for cell in row:
                new_cell = target_sheet.cell(row=cell.row, column=cell.column)
                
                # Copy value
                new_cell.value = cell.value
                
                # Copy style if exists
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()
        
        return target_sheet
    
    def is_cell_in_merged_range(self, sheet, row: int, col: int) -> Tuple[bool, int, int]:
        """
        Check if a cell is part of a merged range.
        
        Returns:
            Tuple of (is_merged, top_row, top_col)
        """
        for merged_range in sheet.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
                return True, merged_range.min_row, merged_range.min_col
        return False, row, col
    
    def find_template_gstr2b(self, input_path: Path, month: str, year: str) -> Optional[Path]:
        """
        Find a suitable GSTR-2B file to use as template.
        
        When creating new GSTR-2B files from IMS, we use an existing
        GSTR-2B as a template to maintain consistent formatting.
        
        Args:
            input_path: Input folder path
            month: Target month
            year: Target year
        
        Returns:
            Path to template file or None
        """
        from utils.file_utils import check_file_month
        
        gstr2b_files = [f for f in input_path.glob("GSTR2B-*.xlsx")
                        if check_file_month(f.stem, month, year)]
        
        if gstr2b_files:
            return gstr2b_files[0]
        return None
    
    def merge_ims_into_gstr2b(self, gstr_file: Path, ims_file: Path, 
                              output_file: Path) -> ProcessingResult:
        """
        Merge IMS sheet from IMS Reco file into GSTR-2B file.
        
        This is the main operation for clients with both files.
        
        Args:
            gstr_file: Path to GSTR-2B file
            ims_file: Path to IMS Reco file
            output_file: Path for output file
        
        Returns:
            ProcessingResult with status
        """
        try:
            self.logger.debug(f"Merging: {gstr_file.name} + {ims_file.name}")
            
            # Load both workbooks
            gstr_wb = load_workbook(gstr_file)
            ims_wb = load_workbook(ims_file)
            
            # Check if IMS sheet exists
            if "IMS" not in ims_wb.sheetnames:
                self.logger.warning(f"IMS sheet not found in {ims_file.name}")
                return ProcessingResult(
                    success=False,
                    message="IMS sheet not found in IMS Reco file",
                    error_code="ims_sheet_missing"
                )
            
            # Copy IMS sheet to GSTR-2B workbook
            self.copy_sheet_with_formatting(ims_wb, "IMS", gstr_wb, "IMS")
            
            # Save the merged file
            gstr_wb.save(output_file)
            
            self.logger.debug(f"Saved merged file: {output_file.name}")
            
            return ProcessingResult(
                success=True,
                message="Successfully merged IMS into GSTR-2B",
                output_path=output_file
            )
            
        except PermissionError:
            self.logger.error(f"Permission denied: {gstr_file.name}")
            return ProcessingResult(
                success=False,
                message="File is locked or permission denied",
                error_code="file_locked"
            )
        except Exception as e:
            self.logger.error(f"Merge error: {str(e)}")
            return ProcessingResult(
                success=False,
                message=f"Processing Error: {str(e)}",
                error_code="general_error"
            )
    
    def create_gstr2b_from_ims(self, ims_file: Path, output_folder: Path,
                               month: str, year: str,
                               template_file: Optional[Path] = None) -> ProcessingResult:
        """
        Create a new GSTR-2B file from IMS Reco file.
        
        Used when client has IMS but no GSTR-2B file.
        
        Args:
            ims_file: Path to IMS Reco file
            output_folder: Folder to save new file
            month: Processing month
            year: Processing year
            template_file: Optional template GSTR-2B file
        
        Returns:
            ProcessingResult with status
        """
        try:
            # Extract client info from IMS filename
            client_name, state = extract_client_info(ims_file.stem, "IMS")
            if not client_name or not state:
                return ProcessingResult(
                    success=False,
                    message="Could not extract client info from IMS filename",
                    error_code="filename_parse_error"
                )
            
            self.logger.debug(f"Creating GSTR-2B for: {client_name}-{state}")
            
            # Create new filename
            gstr2b_filename = f"GSTR2B-{client_name}-{state}-{month} {year}.xlsx"
            output_path = output_folder / gstr2b_filename
            
            # Load IMS file
            ims_wb = load_workbook(ims_file)
            
            if template_file and template_file.exists():
                # Use template file
                new_wb = load_workbook(template_file)
                
                # Clear data from sheets (keep structure and headers)
                for sheet_name in new_wb.sheetnames:
                    if sheet_name.lower() != 'ims':
                        sheet = new_wb[sheet_name]
                        cleared_merged = set()
                        
                        # Clear data rows (keep header row 1)
                        for row in range(2, sheet.max_row + 1):
                            for col in range(1, sheet.max_column + 1):
                                is_merged, top_row, top_col = \
                                    self.is_cell_in_merged_range(sheet, row, col)
                                
                                if is_merged:
                                    merged_key = (top_row, top_col)
                                    if merged_key not in cleared_merged and top_row >= 2:
                                        sheet.cell(row=top_row, column=top_col).value = None
                                        cleared_merged.add(merged_key)
                                else:
                                    sheet.cell(row=row, column=col).value = None
            else:
                # Create basic workbook
                new_wb = Workbook()
                new_wb.remove(new_wb.active)
                
                # Create GSTR-2B sheet with basic headers
                gstr_sheet = new_wb.create_sheet("GSTR-2B")
                headers = ["GSTIN", "Trade Name", "Period", "Invoice Number",
                          "Invoice Date", "Invoice Value", "Taxable Value",
                          "IGST", "CGST", "SGST", "CESS"]
                for col, header in enumerate(headers, 1):
                    gstr_sheet.cell(row=1, column=col, value=header)
            
            # Copy IMS sheet from IMS Reco file
            if "IMS" in ims_wb.sheetnames:
                self.copy_sheet_with_formatting(ims_wb, "IMS", new_wb, "IMS")
            
            # Populate Overview sheet if it exists
            self._populate_overview_sheet(new_wb, ims_wb, client_name, month, year)
            
            # Save the new file
            new_wb.save(output_path)
            
            self.logger.debug(f"Created: {output_path.name}")
            
            return ProcessingResult(
                success=True,
                message="Created GSTR-2B file from IMS Reco",
                output_path=output_path
            )
            
        except Exception as e:
            self.logger.error(f"Creation error: {str(e)}")
            return ProcessingResult(
                success=False,
                message=f"Error creating GSTR-2B: {str(e)}",
                error_code="creation_error"
            )
    
    def _populate_overview_sheet(self, new_wb: Workbook, ims_wb: Workbook,
                                 client_name: str, month: str, year: str) -> None:
        """
        Populate the Overview sheet with client data.
        
        Args:
            new_wb: New workbook being created
            ims_wb: Source IMS workbook
            client_name: Client name
            month: Processing month
            year: Processing year
        """
        if "Overview" not in new_wb.sheetnames:
            return
        
        overview_sheet = new_wb["Overview"]
        
        # Try to get data from IMS Overview
        client_name_display = ""
        client_gstin = ""
        
        if "Overview" in ims_wb.sheetnames:
            ims_overview = ims_wb["Overview"]
            try:
                # Get client name from C5
                if ims_overview.cell(row=5, column=3).value:
                    client_name_display = str(ims_overview.cell(row=5, column=3).value)
                
                # Get GSTIN from C7
                if ims_overview.cell(row=7, column=3).value:
                    client_gstin = str(ims_overview.cell(row=7, column=3).value)
            except:
                pass
        
        # Use filename client name if no display name found
        if not client_name_display:
            client_name_display = client_name or "Client Name Not Available"
        
        # Populate Overview sheet
        try:
            # Labels (Column B)
            overview_sheet.cell(row=2, column=2, value="Goyal Tax Services Private Limited")
            overview_sheet.cell(row=5, column=2, value="Company Name")
            overview_sheet.cell(row=6, column=2, value="Contents")
            overview_sheet.cell(row=7, column=2, value="Company GSTIN")
            overview_sheet.cell(row=8, column=2, value="Period")
            overview_sheet.cell(row=9, column=2, value="Creation Time")
            
            # Data (Column C)
            overview_sheet.cell(row=5, column=3, value=client_name_display)
            overview_sheet.cell(row=6, column=3, value="GSTR-2B Data - Net")
            overview_sheet.cell(row=7, column=3, value=client_gstin)
            overview_sheet.cell(row=8, column=3, value=f"{month} {year}")
            
            # Current datetime
            current_dt = datetime.datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")
            overview_sheet.cell(row=9, column=3, value=current_dt)
            
        except Exception as e:
            # If Overview population fails, continue anyway
            pass
    
    def copy_file(self, source_file: Path, dest_folder: Path) -> ProcessingResult:
        """
        Copy a file to destination folder.
        
        Used when client has only GSTR-2B (no IMS).
        
        Args:
            source_file: Source file path
            dest_folder: Destination folder
        
        Returns:
            ProcessingResult with status
        """
        try:
            import shutil
            dest_path = dest_folder / source_file.name
            shutil.copy2(source_file, dest_path)
            
            return ProcessingResult(
                success=True,
                message="Copied GSTR-2B (no IMS available)",
                output_path=dest_path
            )
        except Exception as e:
            return ProcessingResult(
                success=False,
                message=f"Copy failed: {str(e)}",
                error_code="general_error"
            )
    
    def set_sheet_tab_color(self, sheet, color: str) -> None:
        """
        Set the tab color for a worksheet.
        
        Args:
            sheet: Worksheet object
            color: Hex color code
        """
        sheet.sheet_properties.tabColor = color
